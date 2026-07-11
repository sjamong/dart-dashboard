from __future__ import annotations

import io
from typing import Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

UNIT_DIVISOR = 1_000_000_000_000  # 1조원

# (표시명, 대상 재무제표 구분(sj_div), 후보 계정명들)
METRICS = [
    ("매출액(조원)", ("IS", "CIS"), ("매출액", "수익(매출액)", "영업수익")),
    ("영업이익(조원)", ("IS", "CIS"), ("영업이익", "영업이익(손실)")),
    ("당기순이익(조원)", ("IS", "CIS"), ("당기순이익", "당기순이익(손실)")),
    ("자산총계(조원)", ("BS",), ("자산총계",)),
    ("부채총계(조원)", ("BS",), ("부채총계",)),
    ("자본총계(조원)", ("BS",), ("자본총계",)),
]

YEAR_ROW_FILLS = [
    PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid"),
    PatternFill(start_color="FFF2F5FA", end_color="FFF2F5FA", fill_type="solid"),
]


def to_number(text: Optional[str]) -> Optional[float]:
    if not text:
        return None
    text = text.replace(",", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def to_jo(text: Optional[str]) -> Optional[float]:
    value = to_number(text)
    if value is None:
        return None
    return round(value / UNIT_DIVISOR, 2)


def extract_metrics(items: list[dict]) -> dict:
    result = {}
    for label, sj_divs, names in METRICS:
        found = None
        # 1차: 정확히 일치하는 계정명
        for item in items:
            if item.get("sj_div") in sj_divs and item.get("account_nm", "").strip() in names:
                found = item
                break
        # 2차: 계정명에 후보 단어가 포함
        if found is None:
            for item in items:
                if item.get("sj_div") in sj_divs and any(n in item.get("account_nm", "") for n in names):
                    found = item
                    break
        result[label] = found
    return result


def build_rows(corp_name: str, year: int, items: list[dict]) -> list[dict]:
    metrics = extract_metrics(items)
    periods = [
        ("thstrm_amount", year),
        ("frmtrm_amount", year - 1),
        ("bfefrmtrm_amount", year - 2),
    ]
    rows = []
    for amount_key, yr in periods:
        row = {"회사명": corp_name, "연도": yr}
        for label, _, _ in METRICS:
            item = metrics[label]
            row[label] = to_jo(item.get(amount_key)) if item else None
        rows.append(row)
    return rows


def add_growth_rates(df: pd.DataFrame) -> pd.DataFrame:
    """회사별로 전년 대비 증감률(%) 컬럼을 각 금액 컬럼 바로 뒤에 추가한다."""
    df = df.sort_values(["회사명", "연도"]).reset_index(drop=True)
    metric_labels = [label for label, _, _ in METRICS]
    growth_col_of = {}
    for label in metric_labels:
        base_name = label.split("(")[0]
        growth_col = f"{base_name} 증감률(%)"
        growth_col_of[label] = growth_col
        df[growth_col] = df.groupby("회사명")[label].pct_change().mul(100).round(2)

    ordered = ["회사명", "연도"]
    for label in metric_labels:
        ordered.append(label)
        ordered.append(growth_col_of[label])
    return df[ordered]


def add_ratios(df: pd.DataFrame) -> pd.DataFrame:
    df["영업이익률(%)"] = (df["영업이익(조원)"] / df["매출액(조원)"] * 100).round(2)
    df["순이익률(%)"] = (df["당기순이익(조원)"] / df["매출액(조원)"] * 100).round(2)
    df["부채비율(%)"] = (df["부채총계(조원)"] / df["자본총계(조원)"] * 100).round(2)
    df["ROE(%)"] = (df["당기순이익(조원)"] / df["자본총계(조원)"] * 100).round(2)
    return df


def build_comparison_df(companies: list[tuple[str, list[dict]]], year: int) -> pd.DataFrame:
    """companies: [(회사명, get_financial_statements() 결과), ...]"""
    all_rows: list[dict] = []
    for corp_name, items in companies:
        all_rows.extend(build_rows(corp_name, year, items))
    df = pd.DataFrame(all_rows)
    df = add_growth_rates(df)
    df = add_ratios(df)
    df = df.sort_values(["연도", "회사명"], ascending=[False, True]).reset_index(drop=True)
    return df


def export_excel(df: pd.DataFrame, path_or_buffer) -> None:
    with pd.ExcelWriter(path_or_buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="재무비교", index=False)
        ws = writer.sheets["재무비교"]
        columns = list(df.columns)

        for col_idx, col_name in enumerate(columns, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(12, len(col_name) + 4)

        prev_year = None
        fill_idx = -1
        for i, row in enumerate(df.itertuples(index=False)):
            row_idx = i + 2
            year_value = getattr(row, "연도")
            if year_value != prev_year:
                fill_idx = (fill_idx + 1) % 2
                prev_year = year_value
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = YEAR_ROW_FILLS[fill_idx]
                if "증감률" in col_name:
                    cell.number_format = "+0.00;-0.00;0.00"
                elif col_name.endswith("(%)"):
                    cell.number_format = "0.00"
                elif col_name not in ("회사명", "연도"):
                    cell.number_format = "#,##0.00"

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")


def export_excel_bytes(df: pd.DataFrame) -> bytes:
    """디스크에 쓰지 않고 엑셀 바이트를 메모리로 반환 (서버리스 환경에서 디스크에 의존하지 않기 위함)."""
    buffer = io.BytesIO()
    export_excel(df, buffer)
    return buffer.getvalue()
